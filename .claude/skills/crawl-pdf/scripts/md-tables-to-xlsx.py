#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
md-tables-to-xlsx.py — Tạo bản XLSX người-đọc SONG SONG từ file Markdown chứa bảng.

Mục tiêu: con người đọc được (Excel) trong khi agent vẫn dùng bản .md gốc.
Chạy hoàn toàn ở máy (openpyxl) → KHÔNG tốn token mô hình.

Quy ước:
  - Mỗi tiêu đề `## <Tên>` (hoặc `# <Tên>`) trong .md = 1 sheet (tên cắt 31 ký tự).
  - Các dòng bảng GFM `| a | b |` trong section → các hàng của sheet.
  - Dòng phân cách `|---|---|` bị bỏ.
  - Dọn artifact bản trích thô: ô `NaN` / `NaT` → rỗng; header `Unnamed: N` → rỗng; bỏ hàng rỗng hoàn toàn.
  - Bảng không nằm dưới tiêu đề nào → gom vào sheet "Sheet1".

Dùng:
  python md-tables-to-xlsx.py <input.md> [output.xlsx]
Nếu thiếu output → cùng tên, đuôi .xlsx, cạnh file .md.
"""
import sys, os, re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Thiếu openpyxl: python -m pip install --user openpyxl")

NA = {"nan", "nat", ""}
def clean(cell: str) -> str:
    c = cell.strip()
    if c.lower() in NA:
        return ""
    if re.fullmatch(r"Unnamed:\s*\d+", c):
        return ""
    return c

def split_row(line: str):
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    return [clean(x) for x in s.split("|")]

def is_sep(cells):
    return all(re.fullmatch(r":?-{2,}:?", c.replace(" ", "")) for c in cells if c != "") and any(cells)

def main():
    if len(sys.argv) < 2:
        sys.exit("Dùng: python md-tables-to-xlsx.py <input.md> [output.xlsx]")
    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(src)[0] + ".xlsx"
    with open(src, encoding="utf-8") as f:
        lines = f.read().splitlines()

    sections = []  # (title, [rows])
    cur_title, cur_rows = "Sheet1", []
    in_front = False
    for ln in lines:
        if ln.strip() == "---" and not cur_rows:
            in_front = not in_front
            continue
        if in_front:
            continue
        m = re.match(r"^#{1,6}\s+(.*)$", ln)
        if m:
            if cur_rows:
                sections.append((cur_title, cur_rows)); cur_rows = []
            cur_title = m.group(1).strip()
            continue
        if ln.lstrip().startswith("|"):
            cells = split_row(ln)
            if is_sep(cells):
                continue
            if any(c != "" for c in cells):
                cur_rows.append(cells)
    if cur_rows:
        sections.append((cur_title, cur_rows))

    if not sections:
        sys.exit("Không tìm thấy bảng nào trong file .md.")

    wb = Workbook(); wb.remove(wb.active)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="D9D9D9")
    used = set()
    for title, rows in sections:
        name = re.sub(r"[\\/?*\[\]:]", " ", title)[:31] or "Sheet"
        base = name; i = 1
        while name.lower() in used:
            i += 1; name = f"{base[:28]}_{i}"
        used.add(name.lower())
        ws = wb.create_sheet(name)
        ncol = max((len(r) for r in rows), default=1)
        for ri, r in enumerate(rows, 1):
            for ci in range(ncol):
                val = r[ci] if ci < len(r) else ""
                cell = ws.cell(row=ri, column=ci + 1, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if ri == 1:
                    cell.font = Font(bold=True, color="000000")
                    cell.fill = head_fill
        ws.freeze_panes = "A2"
        for ci in range(1, ncol + 1):
            maxlen = max((len(str(r[ci - 1])) for r in rows if ci - 1 < len(r)), default=8)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(maxlen + 2, 10), 60)
    wb.save(dst)
    print(f"[OK] {dst}  ({len(sections)} sheet)")

if __name__ == "__main__":
    main()
